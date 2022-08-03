import PySimpleGUI as sg
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
import pytesseract
import cv2
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options
from selenium import webdriver
import os
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



# Парсер Авито ---------------------------------

def Avito(keyword, region):

    os.environ['MOZ_HEADLESS'] = '1'


    options1 = Options()
    options1.add_argument('-headless')
    options1.binary_location = r"C:\Program Files\Firefox Developer " \
                               r"Edition\firefox.exe"
    options1.set_preference("general.useragent.override", UserAgent().random)
    options1.set_preference("network.websocket.enabled", False)
    options1.add_argument('window-size=1600x900')
    profile = webdriver.FirefoxProfile()
    profile.set_preference("dom.webdriver.enabled", False)
    profile.set_preference('useAutomationExtension', False)
    profile.update_preferences()

    driver = webdriver.Firefox(options=options1,
                               executable_path=r'C:\geckodriver' \
                                               r'.exe',
                               firefox_profile=profile)
    driver.implicitly_wait(15)

    view_port_height = "var viewPortHeight = Math.max(document.documentElement.clientHeight, window.innerHeight || 0);"
    element_top = "var elementTop = arguments[0].getBoundingClientRect().top;"
    js_function = "window.scrollBy(0, elementTop-(viewPortHeight/2));"
    scroll_into_middle = view_port_height + element_top + js_function

    actions = ActionChains(driver)

    def get_start_page():


        driver.get("https://www.avito.ru/moskva/predlozheniya_uslug")
        print('Подключение к сайту Avito')
        driver.maximize_window()
        sleep(6)


    def search_params(keyword, region):

        # Вводим ключевое слово для поиска
        input_tb = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'input-input-25uCh')))
        driver.execute_script(scroll_into_middle, input_tb)
        input_tb.clear()
        input_tb.send_keys(keyword)
        input_tb.click()
        print('Ввод ключевого слова поиска на сайте Авито -', keyword)
        sleep(2)

        # Выбираем регион поиска
        region_tb = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'main-text-2PaZG')))
        driver.execute_script(scroll_into_middle, region_tb)
        region_tb.click()
        region_tb2 = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'suggest-input-3p8yi')))
        region_tb2.clear()
        region_tb2.send_keys(region)
        sleep(4)
        region_enter = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'popup-buttons-NqjQ3')))

        region_enter.click()
        print('Выбор региона поиска на сайте Авито - ', region)
        sleep(5)


    def get_urls():

        href_list = []
        main_url = driver.current_url

        for i in range(2, 1000):

            try:
                page = driver.page_source
                print(f'Сбор данных, страница {i-1}')
                page_soup = BeautifulSoup(page, 'lxml')
                main_tag = page_soup.find('div', class_='items-items-38oUm')

                href_tag_list = main_tag.find_all('a',
                                                   class_='link-link-39EVK link-design-default-2sPEv title-root-395AQ iva-item-title-1Rmmj title-listRedesign-3RaU2 title-root_maxHeight-3obWc')
                if len(href_tag_list) == 0:
                    print('Конец')
                    break
                for href_tag in href_tag_list:
                    href = href_tag.get('href')
                    href = 'https://www.avito.ru' + href
                    href_list.append(href)

                currentURL = driver.current_url
                new_url = f'{currentURL}&p={i}'
                try:
                    flag_elem = driver.find_elements_by_class_name('items-extraTitle-173_R')
                    if flag_elem:
                        end_flag = False
                    else:
                        end_flag = True
                except Exception:
                    end_flag = True
                if end_flag:
                    try:
                        next_page_ok = driver.find_element_by_class_name(
                            'pagination-root-2oCjZ')
                        driver.get(new_url)
                        sleep(2)
                    except Exception:
                        break
                else:
                    break

            except Exception as ex:
                print(ex)
                break

        print('Найдено', len(href_list))

        sleep(2)
        driver.close()
        return href_list


    def get_data(href_list):
        options1 = Options()
        options1.binary_location = r"C:\Program Files\Firefox Developer " \
                                   r"Edition\firefox.exe"
        options1.set_preference("network.websocket.enabled", False)
        profile = webdriver.FirefoxProfile()
        profile.set_preference("dom.webdriver.enabled", False)
        profile.set_preference('useAutomationExtension', False)
        profile.update_preferences()

        driver = webdriver.Firefox(options=options1,
                                   executable_path=r'C:\geckodriver' \
                                                   r'.exe',
                                   firefox_profile=profile)
        driver.implicitly_wait(20)

        view_port_height = "var viewPortHeight = Math.max(document.documentElement.clientHeight, window.innerHeight || 0);"
        element_top = "var elementTop = arguments[0].getBoundingClientRect().top;"
        js_function = "window.scrollBy(0, elementTop-(viewPortHeight/2));"
        scroll_into_middle = view_port_height + element_top + js_function

        actions = ActionChains(driver)

        data_list = []
        count = 1
        for url in href_list:

            driver.get(url)
            driver.maximize_window()
            sleep(4)

            try:
                name = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR,
                                                    '.item-view-seller-info > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > a:nth-child(1)')))
            except Exception:
                try:
                    name = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR,
                                                        '.item-view-seller-info > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > a:nth-child(1)')))
                except Exception:
                    driver.refresh()
                    try:
                        name = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR,
                                                            '.item-view-seller-info > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > a:nth-child(1)')))
                    except Exception:
                        try:
                            name = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR,
                                                                '.item-view-seller-info > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > a:nth-child(1)')))
                        except Exception:
                            continue

            name = name.text

            try:
                number_open = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR,
                                                    '.button-origin_full-width')))
                number_open.click()
                sleep(2)
                numb = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR,
                                                    '.item-phone-big-number > img:nth-child(1)')))
                numb.screenshot(f'numb{count}.png')
                pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'
                img = cv2.imread(f'numb{count}.png')
                img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                count += 1
                config = r'--oem 3 --psm 6'
                data = pytesseract.image_to_string(img, config='--psm 11')
                data = data.split('\n')
                elem_num = data[0]
                elem_num = elem_num.replace('-', '')
                print(f'{name}: {elem_num} (Авито)')
                data_list.append([name, elem_num])

            except Exception as ex:
                print("Телефон не указан")
                print(ex.args)


        driver.close()
        return data_list



    def get_excel(info_list, keyword, region):
        book = openpyxl.Workbook()
        sheet = book.active
        sheet['A1'] = 'Имя/Наименование'
        sheet['A1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C",
                                       fill_type="solid")
        sheet['B1'] = 'Контакты'
        sheet['B1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C",
                                       fill_type="solid")

        wrap_alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')
        wrap_alignment2 = Alignment(wrap_text=True, horizontal='left',
                                    vertical='center')

        sheet.cell(row=1, column=1).alignment = wrap_alignment
        sheet.cell(row=1, column=2).alignment = wrap_alignment

        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 25

        row_number = 2

        for man in info_list:
            key = man[0]
            value = man[1]
            sheet.cell(row=row_number, column=1).value = key
            sheet.cell(row=row_number, column=1).alignment = wrap_alignment2
            sheet.cell(row=row_number, column=2).value = value
            sheet.cell(row=row_number, column=2).alignment = wrap_alignment2
            row_number += 1

        c = sheet['A2']
        sheet.freeze_panes = c
        book_name = 'АвитоУслуги_' + keyword + '_' + region + '.xlsx'
        book.save(book_name)
        print(f'Все данные с сайта Авито сохранены в файл {book_name}')
        book.close()
        return book_name






    get_start_page()
    search_params(keyword, region)
    href_list = get_urls()
    info_list = get_data(href_list)
    book_name1 = get_excel(info_list, keyword, region)

    return book_name1


# Парсер ЯндексУслуги ------------------------------


def Yandex(keyword, region):

    os.environ['MOZ_HEADLESS'] = '1'

    region = region
    keyword = keyword


    options2 = Options()
    options2.add_argument('-headless')
    options2.binary_location = r"C:\Program Files\Firefox Developer " \
                               r"Edition\firefox.exe"
    options2.set_preference("network.websocket.enabled", False)
    profile2 = webdriver.FirefoxProfile()
    profile2.set_preference("dom.webdriver.enabled", False)

    profile2.set_preference("network.http.pipelining", True)
    profile2.set_preference("network.http.proxy.pipelining", True)
    profile2.set_preference("network.http.pipelining.maxrequests", 8)
    profile2.set_preference("content.notify.interval", 500000)
    profile2.set_preference("content.notify.ontimer", True)
    profile2.set_preference("content.switch.threshold", 250000)
    profile2.set_preference("browser.cache.memory.capacity",
                           65536)  # Increase the cache capacity.
    profile2.set_preference("browser.startup.homepage", "about:blank")
    profile2.set_preference("reader.parse-on-load.enabled",
                           False)  # Disable reader, we won't need that.
    profile2.set_preference("browser.pocket.enabled", False)  # Duck pocket too!
    profile2.set_preference("loop.enabled", False)
    profile2.set_preference("browser.chrome.toolbar_style",
                           1)  # Text on Toolbar instead of icons
    profile2.set_preference("browser.display.show_image_placeholders",
                           False)  # Don't show thumbnails on not loaded images.
    profile2.set_preference("browser.display.use_document_colors",
                           False)  # Don't show document colors.
    profile2.set_preference("browser.display.use_document_fonts",
                           0)  # Don't load document fonts.
    profile2.set_preference("browser.display.use_system_colors",
                           True)  # Use system colors.
    profile2.set_preference("browser.formfill.enable",
                           False)  # Autofill on forms disabled.
    profile2.set_preference("browser.helperApps.deleteTempFileOnExit",
                           True)  # Delete temprorary files.
    profile2.set_preference("browser.shell.checkDefaultBrowser", False)
    profile2.set_preference("browser.startup.homepage", "about:blank")
    profile2.set_preference("browser.startup.page", 0)  # blank
    profile2.set_preference("browser.tabs.forceHide",
                           True)  # Disable tabs, We won't need that.
    profile2.set_preference("browser.urlbar.autoFill",
                           False)  # Disable autofill on URL bar.
    profile2.set_preference("browser.urlbar.autocomplete.enabled",
                           False)  # Disable autocomplete on URL bar.
    profile2.set_preference("browser.urlbar.showPopup",
                           False)  # Disable list of URLs when typing on URL bar.
    profile2.set_preference("browser.urlbar.showSearch",
                           False)  # Disable search bar.
    profile2.set_preference("extensions.checkCompatibility",
                           False)  # Addon update disabled
    profile2.set_preference("extensions.checkUpdateSecurity", False)
    profile2.set_preference("extensions.update.autoUpdateEnabled", False)
    profile2.set_preference("extensions.update.enabled", False)
    profile2.set_preference("general.startup.browser", False)
    profile2.set_preference("plugin.default_plugin_disabled", False)
    profile2.set_preference("permissions.default.image",
                           2)  # Image load disabled again

    profile2.set_preference('dom.ipc.plugins.enabled.libflashplayer.so',
                                  'false')
    profile2.set_preference('useAutomationExtension', False)
    profile2.update_preferences()

    driver2 = webdriver.Firefox(options=options2,
                               executable_path=r'C:\geckodriver' \
                                               r'.exe',
                               firefox_profile=profile2)
    driver2.implicitly_wait(10)

    view_port_height = "var viewPortHeight = Math.max(document.documentElement.clientHeight, window.innerHeight || 0);"
    element_top = "var elementTop = arguments[0].getBoundingClientRect().top;"
    js_function = "window.scrollBy(0, elementTop-(viewPortHeight/2));"
    scroll_into_middle = view_port_height + element_top + js_function

    actions = ActionChains(driver2)
    driver2.get("https://uslugi.yandex.ru/")
    print('Подключение к сайту ЯндексУслуги')
    driver2.maximize_window()
    input_tb = driver2.find_element_by_class_name('textinput__control')
    driver2.execute_script(scroll_into_middle, input_tb)
    input_tb.send_keys(keyword)
    input_tb.send_keys(Keys.ENTER)
    print(f'Ввод ключевого слова поиска на сайте ЯндексУслуги - {keyword}')
    region_tb = driver2.find_element_by_css_selector('.textinput_theme_normal > '
                                         'input:nth-child(1)')
    driver2.execute_script(scroll_into_middle, region_tb)
    region_tb.clear()
    region_tb.send_keys(region)
    region_tb.click()
    print(f'Выбор региона поиска на сайте ЯндексУслуги - {region}')
    sleep(2)
    region_tb.send_keys(Keys.ARROW_DOWN)
    region_tb.send_keys(Keys.ENTER)
    sleep(2)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet['A1'] = 'Имя/Наименование'
    sheet['A1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
    sheet['B1'] = 'Контакты'
    sheet['B1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")

    wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    wrap_alignment2 = Alignment(wrap_text=True, horizontal='left', vertical='center')

    sheet.cell(row=1, column=1).alignment = wrap_alignment
    sheet.cell(row=1, column=2).alignment = wrap_alignment

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 25

    driver2.maximize_window()
    row_number = 2
    for k in range(2, 11):
        try:
            for i in range(4, 25, 2):
                try:
                    sleep(2)

                    tel_button = WebDriverWait(driver2, 7).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, f'div.WorkersListBlendered-WorkerCard:nth-child({i}) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > div:nth-child(2) > div:nth-child(1)'))
                    )

                    view_port_height = "var viewPortHeight = Math.max(document.documentElement.clientHeight, window.innerHeight || 0);"
                    element_top = "var elementTop = arguments[0].getBoundingClientRect().top;"
                    js_function = "window.scrollBy(0, elementTop-(viewPortHeight/2));"
                    scroll_into_middle = view_port_height + element_top + js_function
                    driver2.execute_script(scroll_into_middle, tel_button)

                    tel_button.click()
                    obj_name = WebDriverWait(driver2, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.PhoneLoader-Name')))
                    sleep(2)
                    obj_name = obj_name.text
                    number = WebDriverWait(driver2, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'b.Text_line_xxl')))

                    number = number.text
                    print(obj_name, ':', number, '(ЯндексУслуги)')
                    close_button = WebDriverWait(driver2, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, '.YdoModal-BackButton > span:nth-child(1) > svg:nth-child(1)')))

                    close_button.click()
                    sleep(2)

                    sheet.cell(row=row_number, column=1).value = obj_name
                    sheet.cell(row=row_number, column=1).alignment = wrap_alignment2
                    sheet.cell(row=row_number, column=2).value = number
                    sheet.cell(row=row_number, column=2).alignment = wrap_alignment2
                    row_number += 1
                except Exception:
                    continue
            try:
                next_page = driver2.find_element_by_link_text(str(k))
                driver2.execute_script(scroll_into_middle, next_page)
                next_page.click()
            except Exception:
                print('Поиск окончен')
                break
        except Exception:
            print('Конец поиска')

    c = sheet['A2']
    sheet.freeze_panes = c
    book_name = 'ЯндексУслуги_' + keyword + '_' + region + '.xlsx'
    book.save(book_name)
    print(f'Все данные с сайта ЯндексУслуги сохранены в файл {book_name}')
    book.close()
    sleep(10)
    driver2.close()
    return book_name


# Графический интерфейс ----------------------------

os.environ['MOZ_HEADLESS'] = '1'



form = sg.FlexForm('JobsParser')

my_new_theme = {'BACKGROUND': '#ff0099',
                'TEXT': '#000000',
                'INPUT': '#ffffff',
                'TEXT_INPUT': '#000000',
                'SCROLL': '#c7e78b',
                'BUTTON': ('black', '#ffffff'),
                'PROGRESS': ('#01826B', '#D0D0D0'),
                'BORDER': 1,
                'SLIDER_DEPTH': 0,
                'PROGRESS_DEPTH': 0}

sg.theme_add_new('MyNewTheme', my_new_theme)

sg.theme('My New Theme')

layout = [
    [sg.Text('Введите ключевое слово поиска', font=('Open Sans Light', 11)),
     sg.InputText(size=(35, 10))],
    [sg.Text('                  Введите город поиска', font=('Open Sans Light',
                                                               11)), sg.InputText(size=(35, 10)), sg.Checkbox('Вся '
                                                                  'Россия', checkbox_color='white', font=('Open Sans Light', 11))],
    [sg.Output(size=(88, 10), font=('Open Sans Light', 11))],
    [
     sg.FolderBrowse('Выбор папки', font=('Open Sans Light',
                                                              11)), sg.Text(
        'Выберите папку в которую сохранить результаты поиска',
             font=('Open Sans Light', 11)),],
    [sg.Submit(button_text='Начать поиск', font=('Open Sans Light', 11)),
     sg.Cancel(
        button_text='Закрыть', font=('Open Sans Light', 11))]
]

sg.SetGlobalIcon('logo.ico')

window = sg.Window('SMMagicParser', layout)

while True:
    event, values = window.read()
    direct = values['Выбор папки']
    print(direct)
    if event in (None, 'Закрыть'):
        window.close()
        break
    if event == 'Начать поиск':
        print('Ожидайте окончания поиска данных')
        keyword = values[0]
        if values[2]:
            region = 'Россия'
        else:
            region = values[1]
        print(keyword, region)

        try:
            book_name1 = Yandex(keyword, region)
            book_name2 = Avito(keyword, region)
            print('Сбор данных полностью завершен')
            print(book_name1)
            print(book_name2)
            wb1 = load_workbook(book_name1)
            wb1.save(f'{direct}/{book_name1}')
            wb2 = load_workbook(book_name2)
            wb2.save(f'{direct}/{book_name2}')
        except Exception:
            print('Ошибка при поиске данных, попробуйте снова')
